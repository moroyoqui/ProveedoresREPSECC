"""
Genera el SQL para importar datos de cumplimiento 2026 desde 'matriz para importar.xlsx'.
La salida (stdout) se puede pipar directamente a mysql.
"""

import sys
import os
from datetime import datetime, date
from difflib import SequenceMatcher
import openpyxl

ORGANIZATION_ID = 1
VALIDATED_BY = 1  # admin@repsecc.mx
VALIDATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

COLUMN_TO_DOCTYPE = {
    "SAT":                                          1,
    "IMSS":                                         2,
    "INFONAVIT":                                    3,
    "CONSTANCIA SAT":                               1,   # mismo tipo que SAT, se deduplica
    "VALIDAR EN PADRON EL REPSE":                  12,
    "CONTRATO VIGENTE":                            11,
    "LISTADO DE PERSONAL QUE INGRESA AL PARQUE":   None,
    "NOMINA PDF":                                   9,
    "NOMINA XML":                                   9,   # mismo tipo que NOMINA PDF, se deduplica
    "CEDULA DETERMINACION CUOTAS MENSUAL":         14,
    "CEDULA DETERMINACION CUOTAS BIMESTRAL":       15,
    "SIPARE":                                      16,
    "COMPROBANTE PAGO SEGURIDAD SOCIAL":           17,
}

# Proveedores de la BD (id, legal_name) — hardcodeados para no requerir conexión
DB_SUPPLIERS = [
    (3,  "TRAMESON"),
    (4,  "OBRA Y PROYECTOS OKU"),
    (5,  "Constructora Rono"),
    (6,  "Juan Ruelas"),
    (7,  "CM Electrico"),
    (8,  "Leyra Carolina Valdez Soberanes"),
    (9,  "CONCRETOS ROCAMAR"),
    (10, "ELISA SOFIA LUGO  SANCHEZ"),
    (11, "JOSE ALFREDO SOTELO ESPINOZA"),
]


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def match_supplier(excel_name: str):
    best_score = 0.0
    best = None
    for sid, sname in DB_SUPPLIERS:
        score = similarity(excel_name, sname)
        if score > best_score:
            best_score = score
            best = (sid, sname)
    return best if best_score >= 0.35 else None


def main():
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "matriz para importar.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]

    rows_2026 = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        provider = row[1]
        period_cell = row[5]
        if not provider or not hasattr(period_cell, "year"):
            continue
        if period_cell.year != 2026:
            continue
        rows_2026.append(dict(zip(headers, row)))

    lines = ["-- Importación matriz 2026", "USE repse;", ""]

    for row in rows_2026:
        excel_name = row["PROVEEDOR"].strip()
        period_dt = row["MES QUE CORRESPONDE LA INFORMACION"]
        period_start = f"{period_dt.year}-{period_dt.month:02d}-01"
        repse_folio = row.get("# REGISTRO REPSE")

        match = match_supplier(excel_name)
        if not match:
            lines.append(f"-- SIN MATCH: '{excel_name}' omitido")
            continue

        supplier_id, db_name = match
        lines.append(f"-- Proveedor: '{excel_name}' -> '{db_name}' (id={supplier_id}), periodo={period_start}")

        if repse_folio:
            lines.append(
                f"UPDATE suppliers SET repse_folio = '{repse_folio}' "
                f"WHERE id = {supplier_id} AND (repse_folio IS NULL OR repse_folio = '');"
            )

        seen_dtypes = set()
        for col_name, dtype_id in COLUMN_TO_DOCTYPE.items():
            if dtype_id is None:
                continue
            if dtype_id in seen_dtypes:
                continue
            cell_value = row.get(col_name)
            if cell_value != 2:
                continue

            lines.append(
                f"INSERT IGNORE INTO compliance_cell_validations "
                f"(organization_id, supplier_id, document_type_id, coverage_period_start, "
                f"validated_by, validated_at, created_at, updated_at) VALUES "
                f"({ORGANIZATION_ID}, {supplier_id}, {dtype_id}, '{period_start}', "
                f"{VALIDATED_BY}, '{VALIDATED_AT}', NOW(6), NOW(6));  "
                f"-- {col_name}"
            )
            seen_dtypes.add(dtype_id)

        lines.append("")

    print("\n".join(lines))


if __name__ == "__main__":
    main()
