"""
Importa datos de cumplimiento del año 2026 desde 'matriz para importar.xlsx'.

Lógica:
- Lee todas las filas del Excel donde el año del mes sea 2026.
- Mapea el nombre del proveedor al supplier_id de la BD por similitud de nombre.
- Para cada columna de documento con valor 2, inserta un registro en
  compliance_cell_validations (que pone la casilla en verde en la UI).
- Usa INSERT IGNORE para no duplicar registros ya existentes.
- Actualiza repse_folio en suppliers si el campo está vacío.
"""

import sys
import os
from datetime import datetime, date
from difflib import SequenceMatcher
import openpyxl
import mysql.connector

# ─── Conexión ───────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3307,
    "database": "repse",
    "user": "root",
    "password": "5i5tem@5",
}

ORGANIZATION_ID = 1
VALIDATED_BY = 1  # admin@repsecc.mx
VALIDATED_AT = datetime.now().replace(microsecond=0)

# ─── Mapeo columnas Excel → document_type_id ────────────────────────────────
# Valor 2 en la matriz = documento entregado/cumplido
COLUMN_TO_DOCTYPE = {
    "SAT":                                          1,   # Opinión de cumplimiento SAT (32-D)
    "IMSS":                                         2,   # Opinión de cumplimiento IMSS
    "INFONAVIT":                                    3,   # Opinión de cumplimiento INFONAVIT
    "CONSTANCIA SAT":                               1,   # misma categoría SAT — se ignora si ya existe
    "VALIDAR EN PADRON EL REPSE":                  12,   # REPSE
    "CONTRATO VIGENTE":                            11,   # Contrato
    "LISTADO DE PERSONAL QUE INGRESA AL PARQUE":   None, # sin tipo en catálogo — omitir
    "NOMINA PDF":                                   9,   # CFDI de nómina
    "NOMINA XML":                                   9,   # mismo tipo que PDF — se ignora si ya existe
    "CEDULA DETERMINACION CUOTAS MENSUAL":         14,   # Cedula cuota IMSS
    "CEDULA DETERMINACION CUOTAS BIMESTRAL":       15,   # Cedula de determinacion cuota bimestral
    "SIPARE":                                      16,   # SIPARE
    "COMPROBANTE PAGO SEGURIDAD SOCIAL":           17,   # Comprobante Pago Seguro Social
}

DOCUMENT_COLUMNS = list(COLUMN_TO_DOCTYPE.keys())


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def match_supplier(excel_name: str, db_suppliers: list[tuple]) -> tuple | None:
    """Retorna (id, legal_name) del proveedor más similar."""
    best_score = 0.0
    best = None
    for sid, sname in db_suppliers:
        score = similarity(excel_name, sname)
        if score > best_score:
            best_score = score
            best = (sid, sname)
    return best if best_score >= 0.35 else None


def main():
    # ─── Leer Excel ─────────────────────────────────────────────────────────
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "matriz para importar.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]  # fila 2 = encabezados reales

    rows_2026 = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        provider = row[1]
        period_cell = row[5]
        if not provider or not hasattr(period_cell, "year"):
            continue
        if period_cell.year != 2026:
            continue
        row_dict = dict(zip(headers, row))
        rows_2026.append(row_dict)

    if not rows_2026:
        print("No se encontraron filas del año 2026 en el Excel.")
        sys.exit(0)

    print(f"\nFilas 2026 encontradas: {len(rows_2026)}")
    for r in rows_2026:
        print(f"  • {r['PROVEEDOR']}  —  mes: {r['MES QUE CORRESPONDE LA INFORMACION'].strftime('%Y-%m')}")

    # ─── Conectar a BD ───────────────────────────────────────────────────────
    conn = mysql.connector.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, legal_name FROM suppliers WHERE deleted_at IS NULL")
    db_suppliers = cur.fetchall()

    inserted = 0
    skipped = 0
    not_matched = []

    print()
    for row in rows_2026:
        excel_name = row["PROVEEDOR"].strip()
        period_dt: datetime = row["MES QUE CORRESPONDE LA INFORMACION"]
        period_start = date(period_dt.year, period_dt.month, 1)
        repse_folio = row.get("# REGISTRO REPSE")

        match = match_supplier(excel_name, db_suppliers)
        if not match:
            not_matched.append(excel_name)
            print(f"  [SIN MATCH] '{excel_name}' — se omite")
            continue

        supplier_id, db_name = match
        print(f"\n  Proveedor: '{excel_name}' → '{db_name}' (id={supplier_id}), periodo={period_start}")

        # Actualizar repse_folio si está vacío
        if repse_folio:
            cur.execute(
                "UPDATE suppliers SET repse_folio = %s WHERE id = %s AND (repse_folio IS NULL OR repse_folio = '')",
                (repse_folio, supplier_id),
            )
            if cur.rowcount:
                print(f"    ✓ repse_folio actualizado: {repse_folio}")

        # Insertar validaciones de cumplimiento
        seen_dtypes = set()
        for col_name, dtype_id in COLUMN_TO_DOCTYPE.items():
            if dtype_id is None:
                continue
            if dtype_id in seen_dtypes:
                continue  # evitar duplicado por NOMINA PDF/XML o SAT/CONSTANCIA SAT
            cell_value = row.get(col_name)
            if cell_value != 2:
                continue  # solo valor 2 = cumplido

            try:
                cur.execute(
                    """
                    INSERT IGNORE INTO compliance_cell_validations
                        (organization_id, supplier_id, document_type_id,
                         coverage_period_start, validated_by, validated_at,
                         created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(6), NOW(6))
                    """,
                    (ORGANIZATION_ID, supplier_id, dtype_id, period_start, VALIDATED_BY, VALIDATED_AT),
                )
                if cur.rowcount:
                    inserted += 1
                    seen_dtypes.add(dtype_id)
                    cur.execute("SELECT name FROM document_types WHERE id = %s", (dtype_id,))
                    dtype_name = cur.fetchone()[0]
                    print(f"    ✓ {col_name} → '{dtype_name}' (id={dtype_id})")
                else:
                    skipped += 1
                    seen_dtypes.add(dtype_id)
            except Exception as e:
                print(f"    ERROR en {col_name}: {e}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n{'─'*50}")
    print(f"Registros insertados : {inserted}")
    print(f"Ya existían (ignored): {skipped}")
    if not_matched:
        print(f"Sin match en BD      : {not_matched}")
    print("Importación completada.\n")


if __name__ == "__main__":
    main()
